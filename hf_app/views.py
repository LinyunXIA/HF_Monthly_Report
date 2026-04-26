import os
import json
import threading
from datetime import datetime
from pathlib import Path
from django.shortcuts import render
from django.http import JsonResponse, FileResponse
from django.conf import settings
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment

DATA_DIR = settings.DATA_DIR
DATA_DIR.mkdir(exist_ok=True)

RECORD_FILE = DATA_DIR / 'records.json'

def load_records():
    if RECORD_FILE.exists():
        with open(RECORD_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return []

def save_records(records):
    with open(RECORD_FILE, 'w', encoding='utf-8') as f:
        json.dump(records, f, ensure_ascii=False, indent=2)

def index(request):
    return render(request, 'index.html')

@csrf_exempt
def upload(request):
    print("Upload request received")
    print("FILES:", request.FILES)
    files = request.FILES.getlist('files')
    print("files count:", len(files))
    if not files:
        return JsonResponse({'success': False, 'error': '没有文件'})
    
    records = load_records()
    record_id = len(records) + 1
    now = datetime.now()
    
    record = {
        'id': record_id,
        'upload_time': now.strftime('%Y年%m月%d日%H点%M分%S秒'),
        'file_count': len(files),
        'status': '上传中',
        'output_file': None
    }
    records.append(record)
    save_records(records)
    
    upload_dir = DATA_DIR / f'upload_{record_id}'
    upload_dir.mkdir(exist_ok=True)
    
    for f in files:
        file_path = upload_dir / f.name
        with open(file_path, 'wb') as pf:
            for chunk in f.chunks():
                pf.write(chunk)
    
    record['status'] = '处理中'
    save_records(records)
    
    thread = threading.Thread(target=process_files, args=(record_id, upload_dir))
    thread.start()
    
    return JsonResponse({'success': True, 'record': record})

def process_files(record_id, upload_dir):
    try:
        records = load_records()
        record = next((r for r in records if r['id'] == record_id), None)
        
        if not record:
            return
        
        xlsx_files = list(upload_dir.glob('*.xlsx'))
        
        target_headers = ['事件分类（可选）', '事件描述']
        all_rows = [target_headers]
        
        summary_count = {}
        others_data = []
        
        for xlsx_file in xlsx_files:
            wb = load_workbook(xlsx_file, data_only=True)
            ws = wb.active
            
            rows = list(ws.iter_rows(values_only=True))
            
            if len(rows) < 2:
                continue
            
            file_headers = list(rows[1])
            
            type_idx = None
            desc_idx = None
            for i, h in enumerate(file_headers):
                h_str = str(h) if h else ''
                if '事件分类' in h_str:
                    type_idx = i
                if '事件描述' in h_str:
                    desc_idx = i
            
            if type_idx is None:
                continue
            
            for row in rows[2:]:
                if row and len(row) > type_idx:
                    event_type = row[type_idx]
                    event_desc = row[desc_idx] if desc_idx is not None and desc_idx < len(row) else None
                    
                    if event_type:
                        all_rows.append([event_type, event_desc])
                        
                        event_str = str(event_type).strip()
                        summary_count[event_str] = summary_count.get(event_str, 0) + 1
                        
                        if event_str == '其他':
                            desc = str(event_desc) if event_desc else ''
                            others_data.append([event_str, desc])
        
        if len(all_rows) < 2:
            record['status'] = '错误: 无有效数据'
            save_records(records)
            return
        
        output_file = DATA_DIR / f'result_{record_id}.xlsx'
        
        wb = Workbook()
        
        ws_details = wb.active
        ws_details.title = 'Details'
        for row in all_rows:
            ws_details.append(row)
        
        ws_summary = wb.create_sheet('Summary')
        ws_summary.append(['事件分类（可选）', '数量'])
        for event_type, count in sorted(summary_count.items()):
            ws_summary.append([event_type, count])
        
        ws_others = wb.create_sheet('Others')
        ws_others.append(target_headers)
        for row in others_data:
            ws_others.append(row)
        
        wb.save(output_file)
        
        record['status'] = '处理完毕'
        record['output_file'] = str(output_file)
        
    except Exception as e:
        if record:
            record['status'] = f'错误: {str(e)}'
    
    if record:
        save_records(records)

def history(request):
    records = load_records()
    return JsonResponse({'records': records})

def status(request, record_id):
    records = load_records()
    record = next((r for r in records if r['id'] == record_id), None)
    
    if not record:
        return JsonResponse({'status': '未找到'})
    
    return JsonResponse({'status': record['status']})

def download(request, record_id):
    records = load_records()
    record = next((r for r in records if r['id'] == record_id), None)
    
    if not record or not record.get('output_file'):
        return JsonResponse({'error': '文件不存在'}, status=404)
    
    output_path = Path(record['output_file'])
    if not output_path.exists():
        return JsonResponse({'error': '文件不存在'}, status=404)
    
    return FileResponse(
        open(output_path, 'rb'),
        as_attachment=True,
        filename=f'月度报告_{record_id}.xlsx'
    )